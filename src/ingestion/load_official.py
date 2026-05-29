"""
Load official Excel dataset into Supabase.
Usage: python -m src.ingestion.load_official
"""
import os
import openpyxl
from datetime import date, datetime
from supabase import create_client
from dotenv import load_dotenv

load_dotenv()

XLSX = "data/raw/evento/Evento_Datasets_Sinteticos_Fraude_500_v2.xlsx"

SHEET_MAP = {
    "3_Asegurados": ("asegurados_sinteticos", {
        "ID Asegurado": "id_asegurado",
        "Nombres Asegurado": "nombre_completo",
        "Segmento": "segmento",
        "Ciudad": "ciudad",
        "Antigüedad (años)": "antiguedad_anios",
        "N° Pólizas Activas": "polizas_activas",
        "N° Reclamos Últimos 12 Meses": "reclamos_12_meses",
        "N° Reclamos Histórico Total": "reclamos_historico",
        "Reclamos RC sin Tercero": "reclamos_rc_sin_tercero",
        "Perfil Riesgo Histórico": "perfil_riesgo",
    }),
    "4_Proveedores": ("proveedores", {
        "ID Proveedor": "id_proveedor",
        "Nombre Proveedor": "nombre_proveedor",
        "Tipo": "tipo",
        "Ciudad": "ciudad",
        "N° Siniestros Asociados": "n_siniestros_asociados",
        "En Lista Restrictiva": "en_lista_restrictiva",
        "Motivo Restricción": "motivo_restriccion",
        "Promedio Monto ($)": "promedio_monto",
    }),
    "2_Polizas": ("polizas", {
        "ID Póliza": "id_poliza",
        "ID Asegurado": "id_asegurado",
        "Ramo": "ramo",
        "Fecha Inicio": "fecha_inicio",
        "Fecha Fin": "fecha_fin",
        "Suma Asegurada ($)": "suma_asegurada",
        "Prima Anual ($)": "prima_anual",
        "Canal Venta": "canal_venta",
        "Estado Póliza": "estado_poliza",
    }),
    "1_Siniestros": ("siniestros", {
        "ID Siniestro": "id_siniestro",
        "ID Póliza": "id_poliza",
        "ID Asegurado": "id_asegurado",
        "Ramo": "ramo",
        "Placa Vehículo Asegurado": "placa_vehiculo",
        "Cobertura": "cobertura",
        "Fecha Ocurrencia": "fecha_ocurrencia",
        "Fecha Reporte": "fecha_reporte",
        "Días Ocurr→Reporte": "dias_entre_ocurrencia_reporte",
        "Monto Reclamado ($)": "monto_reclamado",
        "Monto Estimado ($)": "monto_estimado",
        "Monto Pagado ($)": "monto_pagado",
        "Estado": "estado",
        "Sucursal": "sucursal",
        "ID Proveedor": "id_proveedor",
        "Descripción del Evento": "descripcion_hechos",
        "Docs Completos": "documentos_completos",
        "Prov. Lista Restrictiva": "prov_lista_restrictiva",
        "Días desde Inicio Póliza": "dias_desde_inicio_poliza",
        "Días hasta Fin Póliza": "dias_hasta_fin_poliza",
        "N° Reclamos Previos Asegurado": "reclamos_previos_asegurado",
        "Suma Asegurada ($)": "suma_asegurada",
        "Similitud Narrativa Máx.": "similitud_narrativa_max",
        "Número Parte Policial": "numero_parte_policial",
    }),
    "5_Documentos": ("documentos", {
        "ID Documento": "id_documento",
        "ID Siniestro": "id_siniestro",
        "Tipo Documento": "tipo_documento",
        "Nombre Archivo PDF": "nombre_archivo",
    }),
}

DATE_FIELDS = {"fecha_inicio", "fecha_fin", "fecha_ocurrencia", "fecha_reporte"}
SINIESTROS_SKIP_PERSIST = {"score_reglas", "nivel_riesgo", "alertas", "score_anomalia", "embedding_descripcion"}
NUMERIC_FIELDS = {
    "antiguedad_anios", "polizas_activas", "reclamos_12_meses", "reclamos_historico",
    "reclamos_rc_sin_tercero", "n_siniestros_asociados", "promedio_monto",
    "suma_asegurada", "prima_anual", "monto_reclamado", "monto_estimado", "monto_pagado",
    "dias_entre_ocurrencia_reporte", "dias_desde_inicio_poliza", "dias_hasta_fin_poliza",
    "reclamos_previos_asegurado", "similitud_narrativa_max",
}


def coerce_value(key: str, val):
    """Normalize cell values for Supabase insertion."""
    if val is None or val == "" or str(val).strip() in ("—", "-", "N/A", "n/a"):
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat() if isinstance(val, datetime) else str(val)
    if key in DATE_FIELDS and isinstance(val, str):
        try:
            return datetime.strptime(val.strip(), "%Y-%m-%d").date().isoformat()
        except ValueError:
            return str(val).strip()
    if isinstance(val, float):
        if val.is_integer():
            return int(val)
        return val
    if isinstance(val, str):
        stripped = val.strip()
        if stripped == "":
            return None
        # Text landed in numeric column → null it out to avoid type error
        if key in NUMERIC_FIELDS:
            try:
                return int(stripped) if stripped.isdigit() else float(stripped)
            except ValueError:
                print(f"  WARN: non-numeric '{stripped[:60]}' in column '{key}', setting NULL")
                return None
        return stripped
    return val


def load_sheet(ws, col_map: dict) -> list[dict]:
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for excel_col, db_col in col_map.items():
            if db_col in SINIESTROS_SKIP_PERSIST:
                continue
            try:
                idx = headers.index(excel_col)
                record[db_col] = coerce_value(db_col, row[idx])
            except (ValueError, IndexError):
                pass  # column not found in this sheet variant
        if any(v is not None for v in record.values()):
            rows.append(record)
    return rows


def batch_upsert(supabase_client, table: str, rows: list[dict], chunk=100):
    for i in range(0, len(rows), chunk):
        batch = rows[i: i + chunk]
        resp = supabase_client.table(table).upsert(batch, on_conflict="id_siniestro" if table == "siniestros" else None).execute()
        print(f"  [{table}] upserted rows {i}–{i+len(batch)-1}")


def main():
    url = os.environ["SUPABASE_URL"]
    key = os.environ["SUPABASE_ANON_KEY"]
    sb = create_client(url, key)

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    # Load in FK-dependency order
    load_order = ["3_Asegurados", "4_Proveedores", "2_Polizas", "1_Siniestros", "5_Documentos"]

    for sheet_name in load_order:
        if sheet_name not in SHEET_MAP:
            continue
        table, col_map = SHEET_MAP[sheet_name]

        # Find matching sheet (name may differ slightly)
        matched = None
        for s in wb.sheetnames:
            if sheet_name in s or s in sheet_name:
                matched = s
                break
        if not matched:
            print(f"  WARNING: sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[matched]
        rows = load_sheet(ws, col_map)
        print(f"Sheet '{matched}' → {table}: {len(rows)} rows")
        if rows:
            batch_upsert(sb, table, rows)

    print("✅ Load complete.")


if __name__ == "__main__":
    main()
